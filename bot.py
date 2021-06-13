# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import os
import datetime
from datetime import timedelta
from pytz import timezone, utc
import openpyxl
import requests
from bs4 import BeautifulSoup

application=Flask(__name__)

# 참고 사항
# 변수명 앞에 d가 붙은 것은 저장된 데이터에서 불러온 값, d가 붙지 않은 것은 현재 or 입력한 데이터 값

KST=timezone('Asia/Seoul')
Days = ["일요일","월요일","화요일","수요일","목요일","금요일","토요일"] # 요일 이름
mealname = ["아침","점심","저녁"] # 식사 이름
mday = [31,28,31,30,31,30,31,31,30,31,30,31] # 매월 일 수
Msg = [["[오늘 아침]","[오늘 점심]","[오늘 저녁]"],["[내일 아침]","[내일 점심]","[내일 저녁]"]] # 급식 title
Menu = [["","",""],["","",""]] # 오늘, 내일 급식
Menu_saved_date = "" # 급식 불러온 날짜
classn = ["11","12","13","14","21","22","23","24","31","32","33","34"] # 반 이름
classN = [20,20,20,20,20,20,20,19,14,13,10,11] # 반 학생 수
classtime = [["08:20","08:30"],["08:40","09:30"],["9:40","10:30"],["10:40","11:30"],["11:40","12:30"],["13:30","14:20"],["14:30","15:20"],["15:30","16:20"]]
Timetable = [[[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]]],
             [[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]]],
             [[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]]]]
Name = [["곽성은","김도연","김민성","김수민","김정빈","김하진","박서우","박수현","안도윤","오은찬","윤효상","이어진","이주희","이철욱","정승현","정태경","정혜인","채정현","최경호","현준하"],
        ["곽민철","김건우","김건호","김라영","김서한","김수현","김태윤","박서진","박종휘","박한상","배항준","서준희","양희정","이상형","이재덕","임경규","장준영","장혁진","전성빈","조재현"],
        ["김준우","김하준","남민주","박예진","박지훈","박현상","배성렬","서정현","서현","이근영","이유진","이지훈","전재욱","조문경","조현성","조현준","차승빈","최요훈","편예준","허진호"],
        ["강지운","고정우","구현우","권나리","김근형","김민석","류현서","배원호","배준형","배창빈","신수원","염민호","이효준","정민규","정재환","정호원","주선우","하도헌","하승민","황우성"],
        ["고영건","구예찬","김규리","김규태","김동건","김동현","김민경","김상한","김세현","김유진","김정수","김진서","박신후","박주용","백길홍","이동현","이성섭","장재혁","최다민","최동하"],
        ["권도형","김민준","김민지","김세인","김주원","김지원","박재범","박재영","박형준","신은규","유혜원","이동건","이영해","이재훈","이효욱","전서희","전승진","전제빈","정현석","최한솔"],
        ["강민성","권민철","김동민","김동은","김승진","김재헌","노동완","문재영","민수현","박현제","서원준","신재한","용유성","이민주","이슬찬","이준엽","전민경","정은주","정의민","최승빈","최정빈"],
        ["강호연","권인구","김성윤","김재용","김태윤","류현웅","박강민","박근형","박성민","배지민","손연우","은성민","은태호","이규원","이지언","전우주","정성엽","정유라","주윤찬","허도윤"],
        ["","","","","","","","","","","","","",""],
        ["","","","","","","","","","","","",""],
        ["","","","","","","","","",""],
        ["","","","","","","","","","",""]]
#statsdict = {'2104' : (1.0000, 1), '2417' : (1.0000, 1), '2402' : (0.964705882352941, 3), '2201' : (0.964705882352941, 3), '2321' : (0.952941176470588, 5), '2118' : (0.952941176470588, 5), '2216' : (0.952941176470588, 5), '2309' : (0.952941176470588, 5), '2109' : (0.952941176470588, 5), '2119' : (0.952941176470588, 5), '2415' : (0.952941176470588, 5), '2107' : (0.941176470588235, 12), '2316' : (0.929411764705882, 13), '2106' : (0.929411764705882, 13), '2115' : (0.882352941176471, 15), '2414' : (0.882352941176471, 15), '2404' : (0.870588235294118, 17), '2302' : (0.870588235294118, 17), '2311' : (0.858823529411765, 19), '2317' : (0.858823529411765, 19), '2406' : (0.835294117647059, 21), '2418' : (0.835294117647059, 21), '2105' : (0.8000, 23), '2113' : (0.8000, 23), '2112' : (0.788235294117647, 25), '2303' : (0.788235294117647, 25), '2405' : (0.776470588235294, 27), '2410' : (0.764705882352941, 28), '2205' : (0.741176470588235, 29), '2401' : (0.741176470588235, 29), '2306' : (0.741176470588235, 29), '2408' : (0.729411764705882, 32), '2407' : (0.729411764705882, 32), '2217' : (0.705882352941177, 34), '2110' : (0.694117647058824, 35), '2420' : (0.670588235294118, 36), '2116' : (0.670588235294118, 36), '2312' : (0.670588235294118, 36), '2413' : (0.658823529411765, 39), '2219' : (0.658823529411765, 39), '2210' : (0.647058823529412, 41), '2310' : (0.647058823529412, 41), '2203' : (0.623529411764706, 43), '2212' : (0.623529411764706, 43), '2213' : (0.611764705882353, 45), '2206' : (0.611764705882353, 45), '2208' : (0.588235294117647, 47), '2313' : (0.588235294117647, 47), '2314' : (0.588235294117647, 47), '2211' : (0.576470588235294, 50), '2103' : (0.576470588235294, 50), '2319' : (0.505882352941176, 52), '2318' : (0.505882352941176, 52), '2108' : (0.505882352941176, 52), '2214' : (0.494117647058824, 55), '2308' : (0.482352941176471, 56), '2320' : (0.447058823529412, 57), '2412' : (0.411764705882353, 58), '2207' : (0.411764705882353, 58), '2304' : (0.388235294117647, 60), '2220' : (0.376470588235294, 61), '2215' : (0.364705882352941, 62), '2416' : (0.364705882352941, 62), '2202' : (0.341176470588235, 64), '2411' : (0.329411764705882, 65), '2114' : (0.305882352941176, 66), '2305' : (0.305882352941176, 66), '2120' : (0.294117647058824, 68), '2409' : (0.282352941176471, 69), '2102' : (0.270588235294118, 70), '2111' : (0.223529411764706, 71), '2204' : (0.223529411764706, 71), '2301' : (0.2000, 73), '2403' : (0.2000, 73), '2209' : (0.2000, 73), '2117' : (0.2000, 73), '2218' : (0.188235294117647, 77), '2419' : (0.164705882352941, 78), '2101' : (0.141176470588235, 79), '2307' : (0.105882352941176, 80), '2315' : (0.0117647058823529, 81)}

def prin(datas,classN): # 시간표 출력 함수
    
    now=datetime.datetime.utcnow() # 현재 시간
    day=int(utc.localize(now).astimezone(KST).strftime("%w"))
    title=""; answer=""
    subName=datas[0]; subZoomid=datas[1]; subZoompwd=datas[2] # datas: 0=name, 1=zoomid, 2=zoompwd, 3=hangout, 4=classroom, 5=teacher
    subHangout=datas[3]; subClassroom=datas[4]
    trimi=subName.find('(');
    if trimi!=-1: subName=subName[:trimi] # 과목명 출력할 때는 괄호 없애기
    
    if "반" in subName: # 조종례일 경우
        title+=Days[day]+" ["+subName
        if classN==0: title+=" 조례]"
        elif classN==8: title+=" 종례]"
    else : # 기타 과목
        title+=Days[day]+" "+str(classN)+"교시 : ["+subName+"]"
    
    if subName=="동아리": # 링크 첨부
        answer+="동아리 활동 시간입니다."
    if not "none" in subZoomid:
        answer+="줌 : https://zoom.us/j/"+subZoomid+"?pwd="+subZoompwd
    if not "none" in subHangout:
        answer+="행아웃 : https://meet.google.com/lookup/"+subHangout
    if not "none" in subClassroom:
        answer+="\n"+"클래스룸 : https://classroom.google.com/u/0/c/"+subClassroom
    if subName!="동아리" and "none" in subZoomid and "none" in subHangout and "none" in subClassroom:
        answer+="관련 접속 정보가 존재하지 않습니다."
    
    if classN!=8: title+=" ("+classtime[classN][0]+" ~ "+classtime[classN][1]+")" # 종례 제외 시간표시
    return title, answer

@application.route('/link', methods=['POST'])
def response_link(): # 온라인 시간표 대답 함수
    
    now = datetime.datetime.utcnow() # 현재 시간
    day = int(utc.localize(now).astimezone(KST).strftime("%w"))
    hour = int(utc.localize(now).astimezone(KST).strftime("%H"))
    minutes = int(utc.localize(now).astimezone(KST).strftime("%M"))
    classN = 0
    # 기본 : 시작시간 - 20 ~ 종료시간 - 10
    # 예외 : 조례, 종례, 5교시(점심시간)
    if (3 <= hour <= 7) or (hour == 8 and minutes < 23): classN = 0 # 8:20~8:30
    elif (hour == 8 and minutes >= 23) or (hour == 9 and minutes < 20): classN = 1 # 8:40~9:30
    elif (hour == 9 and minutes >= 20) or (hour == 10 and minutes < 20): classN = 2 # 9:40~10:30
    elif (hour == 10 and minutes >= 20) or (hour == 11 and minutes < 20): classN = 3 # 10:40~11:30
    elif (hour == 11 and minutes >= 20) or (hour == 12 and minutes < 20): classN = 4 # 11:40~12:30
    elif (hour == 12 and minutes >= 20) or (hour == 13) or (hour == 14 and minutes < 10): classN = 5 # 13:30~14:20
    elif (hour == 14 and minutes >= 10) or (hour == 15 and minutes < 10): classN = 6 # 14:30~15:20
    elif (hour == 15 and minutes >= 10) or (hour == 16 and minutes < 10): classN = 7 # 15:30~16:20
    elif (hour == 16 and minutes <= 40): classN = 8 # 16:20~16:40
    else : classN = 9 # 수업 없음
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # 학번 불러오기
    lines=fr.readlines()
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1];
        if dusid==userid: stid=dstid
    fr.close()
    if stid=="none":
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "basicCard": {
                            "title": "[학번 등록]",
                            "description": "학번이 등록되어 있지 않습니다.\n아래 버튼을 눌러 학번을 등록해주세요",
                            "buttons": [ { "action": "message", "label": "학번 등록", "messageText": "학번 등록" } ]
                        }
                    }
                ]
            }
        }
    else :
        if day==6 or day==0 :#or classN==9: # 수업이 없는 경우
            res={
                "version": "2.0",
                "template": { "outputs": [ { "simpleText": { "text": "진행 중인 수업이 없습니다." } } ] }
            }
            return jsonify(res)
        else :
            grade=int(stid[0]); classn=int(stid[1])
            items=[]
            fr=open("/home/ubuntu/dg1s_bot/timetable.txt","r")
            lines=fr.readlines()
            for j in range(3):
                for k in range(4):
                    for l in range(5):
                        global Timetable
                        line=lines[26*j+2+6*k+l].split(" ")
                        Timetable[j][k][l]=[]
                        for sub in line:
                            sub=sub.replace("\n","")
                            Timetable[j][k][l].append(sub)
            fr.close()
            for i in range(9): # 해당 요일의 시간표 모두 출력  
                subjectName=Timetable[grade-1][classn-1][day-1][(classN+i)%9]
                fr=open("/home/ubuntu/dg1s_bot/subject data.txt","r")
                lines=fr.readlines()
                isgrade=False # 과목명이 겹칠 경우를 대비해 해당 학년의 과목명이 맞는지 확인하기 위한 변수
                for line in lines:
                    datas=line.split(" ")
                    dname=datas[0];
                    if dname==str(grade)+"학년\n":
                        isgrade=True; continue
                    if isgrade==True and "학년" in dname: break
                    if isgrade==True and dname==subjectName: 
                        title, answer=prin(datas,(classN+i)%9)
                        item={ "title": title, "description": answer }
                        items.append(item)
                        break
                fr.close()
            res={ # 답변
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "carousel": {
                                "type": "basicCard",
                                "items": items
                            }
                        },
                        {
                            "simpleText":{
                                "text": "* 실제 시간표는 위와 다를 수도 있다는 점 유의하시기 바랍니다."
                            }
                        }
                    ]
                }
            }
            return jsonify(res)

def what_is_menu(): # made by 1316, 1301
    
    global Menu, Menu_saved_date
    now = datetime.datetime.utcnow() # 오늘, 내일 날짜
    today = utc.localize(now).astimezone(KST)
    tomorrow = today + timedelta(days=1)
    today_name = " "+str(today.month)+"월 "+str(today.day)+"일 " # 추후 비교용 날짜명 텍스트("_N월_N일_")
    tomorrow_name = " "+str(tomorrow.month)+"월 "+str(tomorrow.day)+"일 "

    if Menu_saved_date == "" or Menu_saved_date != today_name :
        Menu_saved_date = today_name
        Menu = [["","",""],["","",""]]
      
        url = 'https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=%EB%8C%80%EA%B5%AC%EC%9D%BC%EA%B3%BC%ED%95%99%EA%B3%A0%EB%93%B1%ED%95%99%EA%B5%90&oquery=eorndlfrhkrh+rmqtlr&tqi=U%2Ftz5wprvOssslHyxuossssssLN-415573'
        response = requests.get(url) # url로부터 가져오기
        if response.status_code == 200:  
          
            source = response.text # menu_info class 내용 가져오기
            soup = BeautifulSoup(source,'html.parser')
            a = soup.select('.menu_info')
          
            for menu in a:
                menu_text = menu.get_text()
                bracket_i = menu_text.find('[')
                bracket_j = menu_text.find(']')
                menu_day = menu_text[:bracket_i]
                menu_when = menu_text[bracket_i+1:bracket_j]
                menu_content = menu_text[bracket_j+3:].rstrip().replace(" ","\n")
              
                if menu_when == "아침": save_i = 0
                elif menu_when == "점심": save_i = 1
                elif menu_when == "저녁": save_i = 2
              
                if menu_day == today_name: Menu[0][save_i]=menu_content
                elif menu_day == tomorrow_name: Menu[1][save_i]=menu_content
    
    req=request.get_json() # 파라미터 값 불러오기
    askmenu=req["action"]["detailParams"]["ask_menu"]["value"]
    
    now=datetime.datetime.utcnow() # 몇 번째 주인지 계산
    date=int(utc.localize(now).astimezone(KST).strftime("%d"))
    month=int(utc.localize(now).astimezone(KST).strftime("%m"))
    year=int(utc.localize(now).astimezone(KST).strftime("%Y"))
    cday=(year-1)*365+(year-1)//4-(year-1)//100+(year-1)//400
    if (year%4==0 and year%100!=0) or year%400==0: cday+=1
    for i in range(month-1): cday+=mday[i]
    cday+=date
    if askmenu=="내일 급식": cday+=1
    cweek=(cday-1)//7
    cweek-=105407 # 2021-03-02 = 105407번째 주
    classn=["1반","2반","3반","4반"]
    boborder="급식 순서 : "+classn[cweek%4]
    for i in range(1,4): boborder+=' - '+classn[(i+cweek)%4]
    
    hour=int(utc.localize(now).astimezone(KST).strftime("%H")) # Meal 계산
    minu=int(utc.localize(now).astimezone(KST).strftime("%M"))
    if (hour==7 and minu>=30) or (hour>=8 and hour<=12) or (hour==13 and minu<30): Meal="아침" # 가장 최근 식사가 언제인지 자동 계산
    elif (hour==13 and minu>=30) or (hour>=14 and hour<19) or (hour==19 and minu<30): Meal="점심"
    else: Meal="저녁"
    
    i = 0
    
    if Meal == "아침": fi=1; si=2; ti=0 # 아침 점심 저녁 정보 불러오기 및 배열
    elif Meal == "점심": fi=2; si=0; ti=1
    elif Meal == "저녁": fi=0; si=1; ti=2
    if askmenu == "내일 급식": fi=0; si=1; ti=2; i=1
    first = Menu[i][fi]
    second = Menu[i][si]
    third = Menu[i][ti]
    if Menu[i][fi] == "": first = "등록된 급식이 없습니다."
    if Menu[i][si] == "": second = "등록된 급식이 없습니다."
    if Menu[i][ti] == "": third = "등록된 급식이 없습니다."
    return Msg[i][fi], Msg[i][si], Msg[i][ti], first, second, third, boborder

@application.route('/menu', methods=['POST'])
def response_menu(): # 메뉴 대답 함수
    
    msg1, msg2, msg3, menu1, menu2, menu3, boborder = what_is_menu()
    if menu1=="등록된 급식이 없습니다." and menu2=="등록된 급식이 없습니다." and menu3=="등록된 급식이 없습니다.":
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "급식이 없는 날입니다."
                        }
                    }
                ]
            }
        }
    else:
        res={ # 답변
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "carousel": {
                            "type": "basicCard",
                            "items": [
                                { "title": msg1, "description": menu1 },
                                { "title": msg2, "description": menu2 },
                                { "title": msg3, "description": menu3 }
                            ]
                        }
                    }#,
                    #{
                        #"simpleText": {
                             #"text": boborder
                        #}
                    #}
                ]
            }
        }
    return jsonify(res)

@application.route('/colcheck', methods=['POST'])
def check_wp():
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # 학번 불러오기
    lines=fr.readlines()
    fr.close()
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1];
        if dusid==userid: 
            stid=dstid
            break
    
    printmsg=""
    url = 'http://13.209.42.53:5000/colstdata'
    headers = { 'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36'}
    response = requests.get(url,headers=headers) # url로부터 가져오기
    if response.status_code == 200: 
        source = response.text
        lines=source.split("\n")
        for line in lines:
            data=line.rstrip('\n').split(' ')
            if len(data)<4: continue
            datastid=data[0]
            datawarning=data[1]
            datapenalty=data[2]
            datareason=data[3:]
            if stid==datastid:
                printmsg="[경고/벌점 현황]\n학번 : "+stid+"\n경고 "+datawarning+"회, 벌점 "+datapenalty+"점"
                if len(datareason)!=1:
                    reasons=""
                    for reason in datareason:
                        if reason=="none": continue
                        reasons+="\n"+reason.replace('_',' ')[:10]+' '+reason.replace('_',' ')[10:]
                    printmsg+="\n사유 :"+reasons
    
    res={
        "version": "2.0",
        "template": {
            "outputs":[
                {
                    "simpleText": {
                        "text": printmsg
                    }
                }
            ]
        }
    }
    return jsonify(res)

  
@application.route('/seat', methods=['POST'])
def input_seat(): # 좌석 번호 입력 함수
    
    now=datetime.datetime.utcnow()
    Day=int(utc.localize(now).astimezone(KST).strftime("%w"))
    hour=int(utc.localize(now).astimezone(KST).strftime("%H"))
    minu=int(utc.localize(now).astimezone(KST).strftime("%M"))
    if (hour==6 and minu>=50) or (hour>=7 and hour<12) or (hour==12 and minu<30): Meal="아침" # 가장 최근 식사가 언제인지 자동 계산
    elif (hour==12 and minu>=30) or (hour>=13 and hour<18) or (hour==18 and minu<30): Meal="점심"
    else: 
        Meal="저녁"
        if (hour==6 and minu<50) or hour<=5 : Day=(Day+6)%7
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    seat=req["action"]["detailParams"]["table_seat"]["value"]
    p1=req["action"]["detailParams"]["student_id"]["value"] # 같이 앉은 사람
    p2=req["action"]["detailParams"]["student_id1"]["value"] # 같이 앉은 사람
    stid="none"; day=Day; meal=Meal
    
    if seat=='.': seat='X'
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # userdata 저장 및 변경
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1]; dday=datas[2] # data 불러오기
        dmeal=datas[3]; dp1=datas[5]; dp2=datas[6].rstrip('\n')
        if dusid==userid:
            stid=dstid
            if dday!="7": day=int(dday) # 요일
            if dmeal!="none": meal=dmeal # 식사
            if p1=="none" and p2=="none": # 같이 앉은 사람
                p1=dp1; p2=dp2
            elif p1!="none" and p2=="none": # 항상 p1이 p2보다 우선적으로 채워지도록
                if dp1=="none" and dp2=="none": p1=p1; p2=dp2
                elif dp1!="none" and dp2=="none": p2=p1; p1=dp1
                elif dp1!="none" and dp2!="none": p2=p1; p1=dp2
        else : fw.write(line)
    if p2==stid or p2==p1: p2="none" # 입력한 사람이 자기 자신이거나 중복일 경우
    if p1==stid: p1="none"
    fw.write(userid+" "+stid+" "+str(day)+" "+meal+" "+seat+" "+p1+" "+p2+"\n")
    fw.close()
        
    if stid=="none": # 등록 안된 user
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "basicCard": {
                            "title": "[학번 등록]",
                            "description": "학번이 등록되어 있지 않습니다.\n아래 버튼을 눌러 학번을 등록해주세요",
                            "buttons": [ { "action": "message", "label": "학번 등록", "messageText": "학번 등록" } ]
                        }
                    }
                ]
            }
        }
    else:
        stids=stid
        if p1!="none" and p1!=stid: stids+=", "+p1 
        if p2!="none" and p2!=stid and p2!=p1: stids+=", "+p2
        
        quickreplies=[] # 사용자가 기록하지 않은 급식을 찾아서 바로가기 응답 형태로 제공
        checkrecord=[[True,False,False],[False,False,False],[False,False,False],[False,False,False],[False,False,True]]
        fr=open("/home/ubuntu/dg1s_bot/final save.txt","r")
        lines=fr.readlines()
        fr.close()
        for line in lines:
            if line==lines[0]: continue
            if line[:4]==stids[:4] and "none" not in line and (line[5]!='0' and line[5]!='6'): checkrecord[int(line[5])-1][int(line[7])]=True # 기록했으면 True
        for i in range(5):
            if i+1 > Day: break
            for j in range(3):
                if i+1==Day and j>mealname.index(Meal): break
                if checkrecord[i][j]==False: # 현재까지의 급식 중 기록을 하지 않았다면 목록에 추가
                    quickreplies.append({ "action": "block",
                                          "label": Days[i+1]+' '+mealname[j],
                                          "messageText": Days[i+1]+' '+mealname[j]+"으로 변경",
                                          "blockId": "605ee41c6daec409bd3bd43d",
                                          "extra": { "meal": str(i+1)+mealname[j] } })
        quickreplies.reverse() # 최근 급식부터 보여주기 위해 역순
        
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "carousel": {
                            "type": "basicCard",
                            "items": [
                                {
                                    "title": "[저장 확인]",
                                    "description": "학번    "+stids+"\n날짜    "+Days[day]+"\n식사    "+meal+"\n좌석    "+seat,
                                    "buttons": [
                                        { "action": "message", "label": "확인", "messageText": "확인했습니다." },
                                        { "action": "message", "label": "초기화", "messageText": "초기화" }
                                    ]
                                },
                                { 
                                    "thumbnail":{
                                        "imageUrl": "http://k.kakaocdn.net/dn/m2tci/btqOvcSDnnh/STY3XTAYC37ce8RYvulrX0/img_l.jpg", "fixedRatio": "true"
                                    } 
                                }
                            ]
                        }
                    }
                ],
                "quickReplies": quickreplies
            }
        }
    '''yourstats=""
    if stid!="none" and stid[0]=="2": 
        yourstats="[통계 결과]\n"+"기간 : 3월 1주차~4월 4주차(총 8주)\n"+"학번 : "+stid+"\n총 참여율 : "+str(round(statsdict[stid][0]*100,2))+"%\n순위 : "+str(statsdict[stid][1])+"/80"
    res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "[사용 불가]\n"+"지정좌석제로 인해 잠정 중단된 기능입니다."
                        }
                    },
                    {
                        "simpleText": {
                            "text": yourstats
                        }
                    }
                ]
            }
        }'''
    return jsonify(res)

@application.route('/chme', methods=['POST'])
def change_meal(): # 식사 변경 함수
  
    now=datetime.datetime.utcnow()
    Day=int(utc.localize(now).astimezone(KST).strftime("%w"))
    hour=int(utc.localize(now).astimezone(KST).strftime("%H"))
    minu=int(utc.localize(now).astimezone(KST).strftime("%M"))
    if (hour==6 and minu>=50) or (hour>=7 and hour<12) or (hour==12 and minu<30): Meal="아침" # 가장 최근 식사가 언제인지 자동 계산
    elif (hour==12 and minu>=30) or (hour>=13 and hour<18) or (hour==18 and minu<30): Meal="점심"
    else: 
        Meal="저녁"
        if (hour==6 and minu<50) or hour<=5 : Day=(Day+6)%7
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    extra=req["action"]["clientExtra"]["meal"]
    
    day=extra[0]; meal=extra[1:3] # user data에서 식사 변경
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r")
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1]; dseat=datas[4]; dp1=datas[5]; dp2=datas[6].rstrip('\n')
        if dusid==userid:
            stids=dstid
            if dp1!="none" and dp1!=dstid: stids+=", "+dp1 
            if dp2!="none" and dp2!=dstid and dp2!=dp1: stids+=", "+dp2
            seat=dseat
            fw.write(userid+" "+dstid+" "+day+" "+meal+" "+dseat+" "+dp1+" "+dp2+"\n")
        else : fw.write(line)
    fw.close()
    
    quickreplies=[] # 사용자가 기록하지 않은 급식을 찾아서 바로가기 응답 형태로 제공
    checkrecord=[[True,False,False],[False,False,False],[False,False,False],[False,False,False],[False,False,True]]
    fr=open("/home/ubuntu/dg1s_bot/final save.txt","r")
    lines=fr.readlines()
    fr.close()
    for line in lines:
        if line==lines[0]: continue
        if line[:4]==stids[:4] and "none" not in line and (line[5]!='0' and line[5]!='6'): checkrecord[int(line[5])-1][int(line[7])]=True # 기록했으면 True
    for i in range(5):
        if i+1 > Day: break
        for j in range(3):
            if i+1==Day and j>mealname.index(Meal): break
            if checkrecord[i][j]==False: # 현재까지의 급식 중 기록을 하지 않았다면 목록에 추가
                quickreplies.append({ "action": "block",
                                      "label": Days[i+1]+' '+mealname[j],
                                      "messageText": Days[i+1]+' '+mealname[j]+"으로 변경",
                                      "blockId": "605ee41c6daec409bd3bd43d",
                                      "extra": { "meal": str(i+1)+mealname[j] } })
    quickreplies.reverse() # 최근 급식부터 보여주기 위해 역순
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "carousel": {
                        "type": "basicCard",
                        "items": [
                            {
                                "title": "[저장 확인]",
                                "description": "학번    "+stids+"\n날짜    "+Days[int(day)]+"\n식사    "+meal+"\n좌석    "+seat,
                                "buttons": [
                                    { "action": "message", "label": "확인", "messageText": "확인했습니다." },
                                    { "action": "message", "label": "초기화", "messageText": "초기화" }
                                ]
                            },
                            { 
                                "thumbnail":{
                                   "imageUrl": "http://k.kakaocdn.net/dn/m2tci/btqOvcSDnnh/STY3XTAYC37ce8RYvulrX0/img_l.jpg", "fixedRatio": "true"
                                } 
                            }
                       ]
                    }
                }
            ],
            "quickReplies": quickreplies
        }
    }
    return jsonify(res)
   
  
@application.route('/stid', methods=['POST'])
def input_stid(): # 학번 입력 함수
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid=req["action"]["detailParams"]["student_id"]["value"]
    check=False
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # userdata 저장 및 변경
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]
        if dusid==userid: 
            fw.write(userid+" "+stid+" 7 none 0 none none\n")
            check=True
        else : fw.write(line)
    if check==False : fw.write(userid+" "+stid+" 7 none 0 none none\n")
    fw.close()
    res={
        "version": "2.0",
        "template": { "outputs": [ { "simpleText": { "text": "학번이 "+stid+"(으)로 등록되었습니다." } } ] }
    }
    return jsonify(res)

@application.route('/save', methods=['POST'])
def final_save(): # 최종 저장 함수
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # 좌석 저장 후 초기화
    lines=fr.readlines()
    fr.close()
    rw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    fw=open("/home/ubuntu/dg1s_bot/final save.txt","a")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1]; dday=datas[2]; dmeal=datas[3]
        dseat=datas[4]; dp1=datas[5]; dp2=datas[6].rstrip('\n')
        if dusid==userid:
            if dmeal=="아침": dmeal='0'
            elif dmeal=="점심": dmeal='1'
            elif dmeal=="저녁": dmeal='2'
            fw.write(dstid+" "+dday+" "+dmeal+" "+dseat+" -\n")
            if dp1!="none": fw.write(dp1+" "+dday+" "+dmeal+" "+dseat+" *\n")
            if dp2!="none": fw.write(dp2+" "+dday+" "+dmeal+" "+dseat+" *\n")
            rw.write(userid+" "+dstid+" 7 none 0 none none\n")
        else : rw.write(line) 
    rw.close()
    fw.close()
    
    res={
        "version": "2.0",
        "template": { "outputs": [ { "simpleText": { "text": "저장되었습니다." } } ] }
    }
    return jsonify(res)

@application.route('/reset', methods=['POST'])
def reset(): # 초기화
    
    now=datetime.datetime.utcnow() # Meal 계산
    Day=int(utc.localize(now).astimezone(KST).strftime("%w"))
    hour=int(utc.localize(now).astimezone(KST).strftime("%H"))
    minu=int(utc.localize(now).astimezone(KST).strftime("%M"))
    if (hour==6 and minu>=50) or (hour>=7 and hour<12) or (hour==12 and minu<30): Meal="아침" # 가장 최근 식사가 언제인지 자동 계산
    elif (hour==12 and minu>=30) or (hour>=13 and hour<18) or (hour==18 and minu<30): Meal="점심"
    else: 
        Meal="저녁"
        if (hour==6 and minu<50) or hour<=5 : Day=(Day+6)%7
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # 초기화
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0];
        if dusid==userid: stid=datas[1]
        if dusid!=userid: fw.write(line)
    fw.write(userid+" "+stid+" 7 none 0 none none\n")
    fw.close()
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "carousel": {
                        "type": "basicCard",
                        "items": [
                            {
                                "title": "[저장 확인]",
                                "description": "학번    "+stid+"\n날짜    "+Days[Day]+"\n식사    "+Meal+"\n좌석    0",
                                "buttons": [
                                    { "action": "message", "label": "확인", "messageText": "확인했습니다." },
                                    { "action": "message", "label": "초기화", "messageText": "초기화" }
                                ]
                            },
                            { 
                                "thumbnail":{
                                    "imageUrl": "http://k.kakaocdn.net/dn/L689z/btqJ78BkcF5/oG7PgVEcPhCqma4ZwyvwAk/img_l.jpg", "fixedRatio": "true"
                                }
                            }
                        ]
                    }
                }
            ]
        }
    }
    return jsonify(res)

@application.route('/excel', methods=['POST'])
def to_excel(): # 엑셀 파일로 생성
    
    wb = openpyxl.load_workbook('Gbob.xlsx',data_only=True) # 엑셀 기본 형식
    
    fr=open("/home/ubuntu/dg1s_bot/final save.txt","r") # 엑셀 채워 넣기
    lines=fr.readlines()
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(" ")
        if len(datas)!=5: continue
        dstid=datas[0]; dday=int(datas[1]); dmeal=int(datas[2]); dseat=datas[3]
        col=dday*3+dmeal; row=int(dstid[2:])+3 
        if dseat==".": dseat="X"
        if 4<=col and col<=16:
            sheet=wb[dstid[:2]]
            sheet.cell(row,col).value=dseat
    fr.close()
    
    sh = wb['통계']
    j = 0
    for sheet in wb:
        if sheet.title not in classn: continue
        T = sheet.title; N = str(classN[j]+3)
        # 통계 칸 채우기
        sh.cell(2,6).value = int(lines[0].rstrip('\n'))
        sh.cell(j+2,5).value = int(N)-3
        sh.cell(j+2,4).value = "=COUNTA("+T+"!D4:P"+N+")/('통계'!$F$2*("+N+"-3))"
        sh.cell(j+2,4).number_format = "0.00%"
        for k in range(4,4+classN[j]):
            # 참여율 칸 채우기
            K = str(k)
            sheet.cell(k,17).value = "=COUNTA(D"+K+":P"+K+")/'통계'!$F$2"
            sheet.cell(k,17).number_format = "0%"
        j += 1
    
    wb.save("bob.xlsx")
    res={
        "version": "2.0",
        "template": {
            "outputs": [ { "simpleText": { "text": "Excel 파일 생성 완료" } } ]
        }
    }
    return jsonify(res)

@application.route('/upst', methods=['POST'])
def update_stid(): # 학번 갱신 함수
    
    updatestr=""
    # 형식: "이전 학번_새 학번_..." ex) "1301 2106 1316 2417" 
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r")
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        former_stid=line.split(" ")[1]
        i=updatestr.find(former_stid)
        if i!=-1: 
            new_stid=updatestr[i+5:i+9]
            line=line.replace(former_stid,new_stid)
        fw.write(line)
    fw.close()
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [ { "simpleText": { "text": "학번 갱신 완료" } } ]
        }
    }
    return jsonify(res)
 
@application.route('/')
def index():
    return render_template("index.html")

filename=""

@application.route('/texteditor')
def text_editor(): # 원하는 파일 사이트에서 보여주고 편집
    global filename
    filename=request.args.get('filename')
    fr=open("/home/ubuntu/dg1s_bot/"+filename+".txt","r")
    data_send=fr.readlines()
    fr.close()
    if filename=="user data": data_send.sort(key=lambda x:x[13:17]) # 학번 순 정렬
    return render_template("texteditor.html",data=data_send, name=filename)

@application.route('/filesave', methods=['GET','POST'])
def save_as_file(): # txt file 저장하기
    if request.method=='POST':
        fr=open("/home/ubuntu/dg1s_bot/"+filename+".txt","r")
        before=fr.read()
        fr.close()
        
        text=request.form['content']
        text=str(text)
        with open("/home/ubuntu/dg1s_bot/"+filename+".txt","w",encoding='utf-8') as f:
            f.write(text)
    
        now=datetime.datetime.utcnow()
        hour=utc.localize(now).astimezone(KST).strftime("%H")
        minu=utc.localize(now).astimezone(KST).strftime("%M")
        date=utc.localize(now).astimezone(KST).strftime("%d")
        month=utc.localize(now).astimezone(KST).strftime("%m")
        year=utc.localize(now).astimezone(KST).strftime("%Y")
        fw=open("/home/ubuntu/dg1s_bot/log.txt","a")
        fw.write('['+year+'-'+month+'-'+date+' '+hour+':'+minu+"] '"+filename+".txt' saved (Below is the contents before saving.)\n")
        fw.write(before+'\n')
        fw.close()
        
        return render_template("saved.html")
  
@application.route('/xlsave', methods=['GET','POST'])
def save_as_xlfile(): # file 저장하기
    if request.method == 'POST':
        f=request.files['xlfile']
        f.save("/home/ubuntu/dg1s_bot/"+secure_filename(f.filename))
        return render_template("saved.html")
  
@application.route('/dnldfile', methods=['GET','POST'])
def download_file(): # file 다운받기
    if request.method == 'POST':
        filename=request.form['downloadfilename']
        return send_file("/home/ubuntu/dg1s_bot/"+filename, attachment_filename=filename, as_attachment=True)

@application.route('/file')
def upload_n_download():
    files=os.listdir("/home/ubuntu/dg1s_bot")
    folders=[]
    for file in files:
        if not '.' in file: folders.append(file)
    for folder in folders:
        files.remove(folder)
    return render_template("file.html", files=files)

@application.route('/status')
def record_status():
    index=int(request.args.get('index'))
    n=classN[index]
    
    stid=[]
    for i in range(1,classN[index]+1):
        id=classn[index]
        if i<10: id+='0'
        id+=str(i)
        stid.append(id)
    
    name=Name[index]
    
    record=[]
    for i in range(25):
        record.append([])
        for j in range(13):
            record[i].append('')
        record[i].append(0)
    fr=open("/home/ubuntu/dg1s_bot/final save.txt", "r")
    lines=fr.readlines()
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(' '); id=datas[0]; day=int(datas[1]); meal=int(datas[2]); seat=datas[3]
        if id[:2]==classn[index]:
            if 3*day+meal-4<0 or 3*day+meal-4>12: continue
            if record[int(id[2:4])-1][3*day+meal-4]=='': record[int(id[2:4])-1][13]+=1
            if seat==".": seat="X"
            record[int(id[2:4])-1][3*day+meal-4]=seat
    fr.close()
    mealN=int(lines[0].rstrip('\n'))
    for i in range(n):
        record[i][13]=str(round((record[i][13]/mealN)*100))+'%'
    
    return render_template("status.html", n=n, stid=stid, name=name, record=record)

@application.route('/ball')
def ball():
    return render_template("Ball.html")

if __name__ == "__main__":
    application.run(host='0.0.0.0', port=5000)
